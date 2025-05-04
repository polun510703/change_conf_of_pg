import os
import json
import re
import shutil
import pandas as pd
from openpyxl import load_workbook

##
#  Create a unique directory by appending _1, _2, ... to the base path.
##
def make_unique_dir(base_path: str) -> str:
    if not os.path.exists(base_path):
        os.makedirs(base_path, exist_ok=True)
        return base_path

    idx = 1
    while True:
        candidate = f"{base_path}_{idx}"
        if not os.path.exists(candidate):
            os.makedirs(candidate)
            return candidate
        idx += 1

##
#   Regularity of alias:
#     • <table>_partN    → <table>
#     • <table>_pN       → <table>
#     • <table>_<digits> → <table>
#     • Others remain unchanged
##
def get_log_search_alias(alias):
    m = re.match(r"^([A-Za-z0-9_]+?)_(?:part|p)?\d+$", alias)
    if m:
        return m.group(1)
    return alias

##
#   Delete the current day's PostgreSQL log file on the server.
#   Assumes logs follow the pattern: postgresql-<Day>.log
##
def delete_today_log_file(client):
    # Get current day (e.g., Mon, Tue, Wed...)
    stdin, stdout, stderr = client.exec_command("date +%a", get_pty=True)
    day_of_week = stdout.read().decode("utf-8", errors="replace").strip()

    log_path = f"/var/lib/pgsql/data/log/postgresql-{day_of_week}.log"
    delete_cmd = f"rm -f {log_path}"

    print(f"[INFO] Deleting log file: {log_path}")
    stdin, stdout, stderr = client.exec_command(delete_cmd, get_pty=True)
    err = stderr.read().decode("utf-8", errors="replace")
    if err:
        print(f"[WARNING] Error when deleting log file: {err}")
    else:
        print("[INFO] Log file deleted successfully.")

##
#   Recursively traverse the plan tree.
#   If the node contains "Relation Name" and "Alias" fields,
#   it represents a base relation. Add its alias to aliases_set.
##
def get_base_relation_aliases(plan_node, aliases_set):
    if "Relation Name" in plan_node and "Alias" in plan_node:
        aliases_set.add(plan_node["Alias"])

    subplans = plan_node.get("Plans", [])
    for sp in subplans:
        get_base_relation_aliases(sp, aliases_set)

##
#   Read a single .json file (output of EXPLAIN in JSON format),
#   and return all base relation aliases set in the query.
##
def parse_explain_json(file_path):
    with open(file_path, "r", encoding="utf-8") as f:
        data = json.load(f)

    # Determine the structure based on PostgreSQL EXPLAIN JSON format:
    if isinstance(data, list) and len(data) > 0 and isinstance(data[0], dict) and "Plan" in data[0]:
        plan_root = data[0]["Plan"]
    elif isinstance(data, dict) and "Plan" in data:
        plan_root = data["Plan"]
    else:
        print(f"[WARNING] {file_path} does not match the expected EXPLAIN JSON structure, returning an empty set.")
        return set()

    aliases = set()
    get_base_relation_aliases(plan_root, aliases)
    return aliases

##
#   Extract RELOPTINFO (ALIAS) blocks from postgresql tail log file for specified aliases.
#   log_path:   path to the log file
#   aliases:    set of aliases to search for
##
def extract_path_info_from_log(log_path, needed_log_aliases):
    if not os.path.exists(log_path):
        print(f"[ERROR] Log file not found: {log_path}")
        return {}

    with open(log_path, "r", encoding="utf-8", errors="replace") as f:
        all_lines = f.readlines()

    alias_path_map = {a: [] for a in needed_log_aliases}

    current_log_alias = None
    collecting = False

    pattern = re.compile(r'^RELOPTINFO \(([^\)]*)\):')

    for line in all_lines:
        match = pattern.match(line)
        if match:
            collecting = False
            current_log_alias = None

            inside = match.group(1).strip()
            inside_aliases = inside.split()

            for inside_alias in inside_aliases:
                norm_alias = get_log_search_alias(inside_alias)
                if norm_alias in needed_log_aliases:
                    current_log_alias = norm_alias
                    collecting = True
                    alias_path_map[current_log_alias].append(line.rstrip('\n'))
                    break

        elif collecting and current_log_alias:
            alias_path_map[current_log_alias].append(line.rstrip('\n'))

    return alias_path_map

##
#   Output path cost information to a text file and convert it to Excel.
##
def output_path_cost_info(folder_path: str,
                          log_filename: str,
                          keep_in_place: bool = False):
    # If keep_in_place is True, the log file should already be in the folder_path.
    # Otherwise, it will be copied to the new directory.
    log_path = os.path.join(folder_path, log_filename)
    if not os.path.exists(log_path):
        print(f"[ERROR] Unable to find log file: {log_path}")
        return

    json_files = [f for f in os.listdir(folder_path) if f.endswith(".json")]
    if not json_files:
        print(f"[WARNING] No .json files in {folder_path}")
        return

    # Retrieve all aliases from the log file
    all_aliases_from_log = set()
    with open(log_path, "r", encoding="utf-8") as lf:
        for line in lf:
            m = re.match(r"^RELOPTINFO \(([^\)]*)\):", line)
            if m:
                parts = m.group(1).split()
                all_aliases_from_log.update(get_log_search_alias(p) for p in parts)

    for jf in json_files:
        sql_name = os.path.splitext(jf)[0]

        # Determine target directory
        if keep_in_place:
            target_dir = folder_path
        else:
            target_dir = make_unique_dir(os.path.join(folder_path, sql_name))

        # path to the JSON and log files
        json_src  = os.path.join(folder_path, jf)
        json_dst  = os.path.join(target_dir, jf)
        log_dst   = os.path.join(target_dir, f"{sql_name}_{log_filename}")

        if not keep_in_place:
            shutil.copyfile(json_src, json_dst)
            shutil.copyfile(log_path, log_dst)
        else:
            # If keep_in_place, we need to use the original log file and the JSON file in the same folder
            json_dst = json_src
            log_dst  = os.path.join(folder_path, log_filename)

        # Analyze the JSON file
        aliases = parse_explain_json(json_dst) 
        needed = {get_log_search_alias(a) for a in aliases}
        
        # debug message
        # print("[DEBUG] needed =", needed)
        
        alias_path_map = extract_path_info_from_log(log_dst, needed)

        # Output path cost information to a text file
        txt_path = os.path.join(target_dir, f"{sql_name}_pathcost.txt")
        with open(txt_path, "w", encoding="utf-8") as outf:
            # Only write the aliases that are in the log file
            for parent in sorted(alias_path_map):
                lines = alias_path_map[parent]
                if not lines:
                    continue
                outf.write(f"===== RELOPTINFO for alias: {parent} =====\n")
                outf.writelines(l + "\n" for l in lines)
                outf.write("\n")

        # Convert the text file to Excel
        xlsx_path = os.path.join(
            target_dir, f"{sql_name}_path_cost_info.xlsx")
        convert_pathcost_file_to_excel(txt_path, xlsx_path)

        print(f"[INFO] Excel file has been saved to: {xlsx_path}")

##
#   Extract partition ID from the path cost information
##
def _partition_id(path):#
    PART_RE = re.compile(r"_(?:part|p)?(\d+)(?:_|$)", re.IGNORECASE)
    m = PART_RE.search(path.get("Index Name") or "")
    return m.group(1) if m else None

##
#   Insert blank lines between partitions in the path cost information
##
def insert_blank_between_partitions(paths):
    # Remove duplicates
    uniq, seen = [], set()
    for p in paths:
        key = (p["Scan Type"], p["Index Name"],
               p["Rows"], p["Startup Cost"], p["Total Cost"])
        if key not in seen:
            seen.add(key)
            uniq.append(p)

    result     = []
    last_part  = None

    for i, p in enumerate(uniq):
        this_part = _partition_id(p)

        # If this_part is None and the previous one is not, set it to the last one
        if this_part is None and p["Scan Type"] == "SeqScan":
            if i + 1 < len(uniq):
                this_part = _partition_id(uniq[i + 1]) or last_part
            else:
                this_part = last_part

        # Determine if we need to insert a blank line
        if last_part is not None and this_part != last_part:
            result.append({
                "Scan Type": None, "Index Name": None,
                "Rows": None, "Startup Cost": None, "Total Cost": None
            })

        result.append(p)
        last_part = this_part

    return result

##
#   Convert the path cost information text file to an Excel file.
##
def convert_pathcost_file_to_excel(input_file, output_excel):
    with open(input_file, "r", encoding="utf-8") as f:
        lines = f.readlines()

    reloptinfo_pattern = re.compile(r"^===== RELOPTINFO for alias: (.+?) =====$")
    path_list_pattern = re.compile(r"path list:")
    partial_path_list_pattern = re.compile(r"partial path list:")
    path_pattern = re.compile(
        r"^(IdxScan|SeqScan)\((.+?)\).*rows=(\d+).*cost=\s*([\d\.]+)\s*\.\.\s*([\d\.]+)"
    )
    index_name_pattern = re.compile(r"index name:\s*(\S+)")
    required_outer_pattern = re.compile(r"required_outer\s*\(([^)]+)\)")
    
    sub_reloptinfo_pattern     = re.compile(r"^RELOPTINFO \((.+?)\):")
    inside_base_relopt         = True   # If inside base reloptinfo, we can collect path info
    
    cheapest_path_heading = re.compile(r"^cheapest .*path[s]?:", re.IGNORECASE)


    data = {}
    current_alias = None
    current_list_type = None
    last_path = None

    for line in lines:
        line = line.strip()
        if not line:
            continue

        reloptinfo_match = reloptinfo_pattern.match(line)
        if reloptinfo_match:
            alias = reloptinfo_match.group(1)
            if alias.startswith("TABLE_ITEMS_"):
                current_alias = "TABLE_ITEMS"
            else:
                current_alias = alias
                
            inside_base_relopt = True

            if current_alias not in data:
                data[current_alias] = {
                    "path_list": [],
                    "partial_path_list": [],
                    "parameterized_path_list": []
                }
            current_list_type = None
            last_path = None
            continue
        
        sub_rel_match = sub_reloptinfo_pattern.match(line)
        if sub_rel_match:
            inner_alias = sub_rel_match.group(1)

            # When there is a space inside the parentheses, it indicates a join RelOptInfo
            inside_base_relopt = (inner_alias == current_alias and " " not in inner_alias)

            # No matter if it's a join or not, reset the state for the next RELOPTINFO ()
            current_list_type = None
            last_path = None
            continue

        if path_list_pattern.match(line):
            if not inside_base_relopt:  # If not inside base reloptinfo, skip this line
                continue
            current_list_type = "path_list"
            last_path = None
            continue

        if partial_path_list_pattern.match(line):
            if not inside_base_relopt:
                continue
            current_list_type = "partial_path_list"
            last_path = None
            continue
        
        # If we encounter a line that indicates the cheapest path, we stop collecting paths
        if cheapest_path_heading.match(line):
            current_list_type = None
            last_path = None
            continue

        path_match = path_pattern.match(line)
        if path_match and current_alias and inside_base_relopt:
            scan_type, table, rows, startup_cost, total_cost = path_match.groups()
            ro_match = required_outer_pattern.search(line)
            required_outer = ro_match.group(1) if ro_match else None
            try:
                rows = int(rows)
                startup_cost = float(startup_cost)
                total_cost = float(total_cost)
            except ValueError:
                print(f"[ERROR] Skipping invalid cost values: {startup_cost}..{total_cost}")
                continue

            if current_alias == "TABLE_ITEMS" and table != "TABLE_ITEMS":
                continue

            path = {
                "Scan Type": scan_type,
                "Index Name": None,
                "Rows": rows,
                "Startup Cost": startup_cost,
                "Total Cost": total_cost,
                "Required Outer": required_outer,
            }
            
            if required_outer:
                list_key = "parameterized_path_list"
            else:
                list_key = current_list_type
            
            if list_key is None:
                continue            # No path list type found, skip this line

            data[current_alias][list_key].append(path)
            last_path = path
            continue

        index_name_match = index_name_pattern.match(line)
        if index_name_match and last_path:
            index_name = index_name_match.group(1)
            last_path["Index Name"] = index_name
            continue

    with pd.ExcelWriter(output_excel) as writer:
        for alias, lists in data.items():
            sheet_data = [[f"Table: {alias}"]]

            df_path_list = pd.DataFrame(insert_blank_between_partitions(lists["path_list"]))
            df_partial_path_list = pd.DataFrame(insert_blank_between_partitions(lists["partial_path_list"]))
            df_parameterized_path_list = pd.DataFrame(insert_blank_between_partitions(lists["parameterized_path_list"]))

            if not df_path_list.empty:
                sheet_data.append(["Path List:"])
                sheet_data.append(["Scan Type", "Index Name", "Rows", "Startup Cost", "Total Cost"])
                sheet_data.extend(df_path_list.drop(columns=["Required Outer"]).values.tolist())

            sheet_data.append([])

            if not df_partial_path_list.empty:
                sheet_data.append(["Partial Path List:"])
                sheet_data.append(["Scan Type", "Index Name", "Rows", "Startup Cost", "Total Cost"])
                sheet_data.extend(df_partial_path_list.drop(columns=["Required Outer"]).values.tolist())
                
            sheet_data.append([])
                
            if not df_parameterized_path_list.empty:
                sheet_data.append(["Parameterized Path List:"])
                sheet_data.append(["Scan Type", "Index Name", "Rows", "Startup Cost", "Total Cost", "Required Outer"])
                sheet_data.extend(df_parameterized_path_list.values.tolist())

            pd.DataFrame(sheet_data).to_excel(writer, index=False, header=False, sheet_name=alias)

    wb = load_workbook(output_excel)
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        column_widths = [12, 40, 10, 15, 15, 25]

        for i, width in enumerate(column_widths, start=1):
            col_letter = ws.cell(row=1, column=i).column_letter
            ws.column_dimensions[col_letter].width = width

    wb.save(output_excel)